import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from pathlib import Path
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# ---------- Бизнес-логика ----------
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
}

def classify_error(response_or_exception, treat_redirect_as_error):
    if isinstance(response_or_exception, requests.Response):
        status = response_or_exception.status_code
        if status < 400:
            if treat_redirect_as_error and status in (301, 302, 303, 307, 308):
                loc = response_or_exception.headers.get('Location', 'unknown')
                return f"Redirect {status} → {loc}"
            return None
        else:
            return f"HTTP {status} {response_or_exception.reason}"
    else:
        return f"Connection Error: {response_or_exception}"

def check_url(excel_row, url, timeout=8, treat_redirect_as_error=False):
    def _do_request(method):
        try:
            if method == 'HEAD':
                return requests.head(str(url), headers=HEADERS, timeout=timeout, allow_redirects=False)
            else:
                with requests.get(str(url), headers=HEADERS, timeout=timeout, stream=True, allow_redirects=False) as resp:
                    resp.close()
                    return resp
        except Exception as e:
            return e

    first_resp = _do_request('HEAD')
    if isinstance(first_resp, requests.Response) and first_resp.status_code in (403, 405, 501):
        second_resp = _do_request('GET')
        resp = second_resp
    else:
        resp = first_resp

    if isinstance(resp, Exception):
        return (excel_row, url, f"Connection Error: {resp}")

    if resp.status_code < 400:
        if treat_redirect_as_error and resp.status_code in (301, 302, 303, 307, 308):
            loc = resp.headers.get('Location', 'unknown')
            return (excel_row, url, f"Redirect {resp.status_code} → {loc}")
        if 300 <= resp.status_code < 400 and not treat_redirect_as_error:
            try:
                final_resp = requests.head(str(url), headers=HEADERS, timeout=timeout, allow_redirects=True)
                if final_resp.status_code < 400:
                    return (excel_row, url, None)
                else:
                    return (excel_row, url, f"HTTP {final_resp.status_code} {final_resp.reason}")
            except Exception as e:
                return (excel_row, url, f"Connection Error after redirect: {e}")
        return (excel_row, url, None)
    else:
        reason = resp.reason if hasattr(resp, 'reason') else ''
        return (excel_row, url, f"HTTP {resp.status_code} {reason}")

def run_checks(url_info_list, treat_redirect_as_error, progress_callback, done_callback):
    results = []
    total = len(url_info_list)
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(check_url, row, url, 8, treat_redirect_as_error): (row, url)
                   for row, url in url_info_list}
        for i, future in enumerate(as_completed(futures), 1):
            res = future.result()
            if res[2] is not None:
                results.append(res)
            progress_callback(i / total * 100)
    done_callback(results)

# ---------- Графический интерфейс ----------
class URLCheckerApp:
    def __init__(self, root):
        self.root = root
        root.title("Проверка URL из Excel")
        root.geometry("520x420")
        root.resizable(False, False)

        self.file_path = tk.StringVar()
        self.sheet_var = tk.StringVar(value="")
        self.column_var = tk.StringVar()
        self.redirect_var = tk.BooleanVar(value=False)

        self.df = None
        self.wb = None

        # --- Виджеты ---
        tk.Label(root, text="1. Выберите Excel-файл", font=("Arial", 10, "bold")).pack(pady=(15,5))
        frame_file = tk.Frame(root)
        frame_file.pack()
        tk.Entry(frame_file, textvariable=self.file_path, width=40).pack(side=tk.LEFT)
        tk.Button(frame_file, text="Обзор", command=self.choose_file).pack(side=tk.LEFT, padx=5)

        tk.Label(root, text="2. Выберите лист", font=("Arial", 10, "bold")).pack(pady=(15,5))
        self.sheet_menu = ttk.Combobox(root, textvariable=self.sheet_var, state="readonly", width=37)
        self.sheet_menu.pack()
        self.sheet_menu.bind('<<ComboboxSelected>>', self.on_sheet_selected)

        tk.Label(root, text="3. Выберите столбец с URL", font=("Arial", 10, "bold")).pack(pady=(15,5))
        self.column_menu = ttk.Combobox(root, textvariable=self.column_var, state="readonly", width=37)
        self.column_menu.pack()

        options_frame = tk.LabelFrame(root, text="Настройки проверки", padx=10, pady=10)
        options_frame.pack(pady=10, fill="x", padx=20)
        tk.Checkbutton(options_frame, text="Считать редиректы (301, 302, ...) ошибкой",
                       variable=self.redirect_var).pack(anchor="w")

        # Кнопки Проверить и Инструкция
        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=10)
        self.check_btn = tk.Button(btn_frame, text="4. Проверить URL", command=self.start_check,
                                   bg="#4CAF50", fg="black", height=2, width=20)
        self.check_btn.pack(side=tk.LEFT, padx=5)
        self.info_btn = tk.Button(btn_frame, text="Инструкция", command=self.show_instructions,
                                  bg="#2196F3", fg="black", height=2, width=14)
        self.info_btn.pack(side=tk.LEFT, padx=5)

        self.progress = ttk.Progressbar(root, orient="horizontal", length=450, mode="determinate")
        self.progress.pack()

        self.status_label = tk.Label(root, text="")
        self.status_label.pack(pady=5)

    def choose_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        self.file_path.set(path)
        try:
            xl = pd.ExcelFile(path)
            sheets = xl.sheet_names
            self.sheet_menu['values'] = sheets
            if sheets:
                self.sheet_var.set(sheets[0])
                self.load_sheet_columns()
        except Exception as e:
            messagebox.showerror("Ошибка чтения файла", str(e))

    def on_sheet_selected(self, event=None):
        self.load_sheet_columns()

    def load_sheet_columns(self):
        if not self.file_path.get() or not self.sheet_var.get():
            return
        try:
            self.df = pd.read_excel(self.file_path.get(), sheet_name=self.sheet_var.get(), nrows=1)
            cols = list(self.df.columns)
            self.column_menu['values'] = cols
            if cols:
                self.column_var.set(cols[0])
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось прочитать столбцы:\n{e}")

    def start_check(self):
        if not self.file_path.get():
            messagebox.showwarning("Выберите файл", "Сначала выберите Excel-файл")
            return
        sheet = self.sheet_var.get()
        column = self.column_var.get()
        if not sheet or not column:
            messagebox.showwarning("Заполните поля", "Выберите лист и столбец")
            return

        try:
            self.wb = openpyxl.load_workbook(self.file_path.get())
            ws = self.wb[sheet]
        except Exception as e:
            messagebox.showerror("Ошибка чтения файла", f"Не удалось открыть файл через openpyxl:\n{e}")
            return

        try:
            col_idx = self.df.columns.get_loc(column)
            col_letter = get_column_letter(col_idx + 1)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось определить столбец:\n{e}")
            return

        url_info = []
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.hyperlink and cell.hyperlink.target:
                url = cell.hyperlink.target
            else:
                val = cell.value
                if val is None:
                    continue
                url = str(val).strip()
                if not url:
                    continue
            url_info.append((row, url))

        if not url_info:
            messagebox.showinfo("Нет данных", "В выбранном столбце нет URL для проверки")
            return

        self.check_btn.config(state="disabled")
        self.progress['value'] = 0
        self.status_label.config(text="Проверка...")

        treat_redirect = self.redirect_var.get()

        thread = threading.Thread(target=run_checks, args=(
            url_info,
            treat_redirect,
            self.update_progress,
            self.on_check_finished
        ))
        thread.daemon = True
        thread.start()

    def update_progress(self, percent):
        self.progress['value'] = percent
        self.root.update_idletasks()

    def on_check_finished(self, errors):
        self.check_btn.config(state="normal")
        self.progress['value'] = 100
        self.status_label.config(text="Готово")

        if not errors:
            messagebox.showinfo("Отлично!", "Все URL работают корректно.")
            return

        df_errors = pd.DataFrame(errors, columns=['Строка в Excel', 'URL', 'Тип ошибки'])

        # Сохраняем в папку исходного файла
        source_dir = Path(self.file_path.get()).parent
        report_path = source_dir / "report_bad_urls.xlsx"
        df_errors.to_excel(report_path, index=False)

        # Стилизация
        try:
            wb_out = openpyxl.load_workbook(report_path)
            ws_out = wb_out.active
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            for row in ws_out.iter_rows(min_row=2, max_row=ws_out.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.fill = red_fill
            wb_out.save(report_path)
        except Exception:
            pass

        messagebox.showinfo("Отчёт сохранён",
                            f"Найдено {len(errors)} проблемных URL.\nФайл: {report_path}")

    def show_instructions(self):
        """Показывает окно с инструкцией."""
        instr_text = """ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ

1. Нажмите «Обзор» и выберите Excel-файл (расширение .xlsx – ОБЯЗАТЕЛЬНО, старые .xls не поддерживаются).
2. Если в книге несколько листов, выберите нужный лист из выпадающего списка.
3. Укажите столбец, в котором находятся URL.
   Поддерживаются:
   – явные ссылки (текст ячейки начинается с http:// или https://);
   – ячейки с гиперссылками (когда текст скрыт, а ссылка вставлена через Ctrl+K).
4. При необходимости отметьте опцию «Считать редиректы ошибкой» – тогда любой редирект (301, 302 и т.д.) будет отражён в отчёте как ошибка с указанием адреса перехода.
5. Нажмите «Проверить URL».
6. Дождитесь завершения. Если будут найдены проблемные URL, отчёт сохранится в ту же папку, где лежит исходный файл, с именем report_bad_urls.xlsx.

ОГРАНИЧЕНИЯ:
– Файл должен быть формата .xlsx.
– В столбце не должно быть картинок или сложных формул – только текст или гиперссылки.
– Проверка идёт по HTTP/HTTPS. Локальные файлы (file://) не проверяются.
– Для надёжности используется эмуляция браузера, но некоторые сайты всё равно могут блокировать запросы (ошибка 403). В таких случаях URL будет помечен как ошибочный.
"""
        top = tk.Toplevel(self.root)
        top.title("Инструкция")
        top.geometry("600x500")
        text_widget = tk.Text(top, wrap="word", padx=15, pady=15)
        text_widget.insert("1.0", instr_text)
        text_widget.config(state="disabled")
        text_widget.pack(fill="both", expand=True)
        tk.Button(top, text="Закрыть", command=top.destroy).pack(pady=10)

if __name__ == "__main__":
    root = tk.Tk()
    app = URLCheckerApp(root)
    root.mainloop()
